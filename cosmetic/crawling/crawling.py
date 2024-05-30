'''
화장품 커뮤니티 앱 "glowpick"에서 카테고리(스킨, 로션)를 설정해서 크롤링하는 코드
'''

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time

import openpyxl as op
# from bs4 import BeautifulSoup

# WORKBOOK_PATH = "./cream_data.xlsx"
# WORKBOOK_PATH = "./water_cream.xlsx"
WORKBOOK_PATH = "./serum_data.xlsx"

# 드라이버 설정
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

# excel 워크북 설정
wb = op.load_workbook(WORKBOOK_PATH)

# 워크 시트 설정
# ws = wb["스킨토너"]
ws = wb.active

# URL 설정 https://www.glowpick.com/categories/4
url = "https://www.glowpick.com/searches/keywords"
driver.get(url)
time.sleep(5)

# 팝업 배너 닫기
try:
    pop_up_close = driver.find_element(By.CLASS_NAME,
                                       'buttons__button.buttons__close')
    pop_up_close.click()
except Exception as e:
    print("팝업 배너가 없거나 닫는 데 실패했습니다:", e)

# 검색 상자 찾기 및 입력
search_box = driver.find_element(By.XPATH,
                                 '//*[@id="default-layout"]/div/div[2]/div/div/div/input')
search_box.send_keys("에센스/세럼")
search_box.send_keys(Keys.RETURN)
time.sleep(2)

# "제품" 카테고리 버튼 클릭
category_button = driver.find_element(By.XPATH,
                                      '//*[@id="contents"]/section/div[1]/nav/ul/li[3]/span')
category_button.click()
time.sleep(3)

# 제품 리스트 추출
products = driver.find_elements(By.CLASS_NAME, 'products__product.product')
print(len(products))

# 각 제품 클릭 및 정보 추출
for idx in range(400):
    data_dict = {
        "category": "에센스/세럼",
        "brand_name": "",
        "product_name": "",
        "price": "",
        "rate": "",
        "number_reviews": "",
        "ingredients": [],
        "descriptions": "",
        "tags": []
    }
    if idx % 4 == 0 and idx != 0:
        # 스크롤 내리기 (2칸 정도만)
        scroll_step = 400
        scroll_pause_time = 2

        driver.execute_script(f"window.scrollBy(0, {scroll_step});")
        time.sleep(scroll_pause_time)

    # 제품 리스트 새로고침
    products = driver.find_elements(By.CLASS_NAME, 'products__product.product')
    print(len(products))
    product = products[idx]
    print(idx)

    # 제품 클릭
    product.click()
    time.sleep(10)  # 페이지 로딩 대기

    # 제품 정보 추출
    brand_name = driver.find_element(By.CLASS_NAME,
                                     'product__summary__brand__name')
    product_name = driver.find_element(By.CLASS_NAME,
                                       'product__summary__name')
    product_price = driver.find_element(By.CLASS_NAME,
                                        'offer__volume-price.font-spoqa')
    product_rating = driver.find_element(By.CLASS_NAME,
                                         'stars__rating')
    number_reviews = driver.find_element(By.CLASS_NAME,
                                         'offer__reviews')

    # dict 추가
    data_dict["brand_name"] = brand_name.text
    data_dict["product_name"] = product_name.text
    data_dict["price"] = product_price.text
    data_dict["rate"] = product_rating.text
    data_dict["number_reviews"] = number_reviews.text

    # 스크롤 내리기
    scroll_step = 150
    scroll_pause_time = 2
    driver.execute_script(f"window.scrollBy(0, {scroll_step});")
    time.sleep(scroll_pause_time)

    # 제품 성분 상세보기 클릭 후 성분 이름 추출
    product_ingre_button = driver.find_element(By.CSS_SELECTOR,
                                               '#contents > section > div.product__info.info > article.info__article.ingredient > h2 > button')
    product_ingre_button.click()
    time.sleep(2)
    product_ingredients = driver.find_elements(By.CLASS_NAME,
                                               'item__wrapper__text__kor')

    data_dict["ingredients"] = [ingre.text for ingre in product_ingredients]
    # print(f"Ingre_list : {[ingre.text for ingre in product_ingredients]}")

    # 상세보기 페이지 닫기
    x_button = driver.find_element(By.CLASS_NAME, 'icon.x-black')
    x_button.click()
    time.sleep(2)

    # 제품 설명 및 태그 추출
    product_description = driver.find_element(By.CLASS_NAME,
                                              'info__text.description__text')

    # tag 추출
    product_tags = driver.find_elements(By.CLASS_NAME,
                                        'description__list__item.info__text.info__text-gray')
    time.sleep(2)

    data_dict["descriptions"] = product_description.text
    data_dict["tags"] = [tag.text for tag in product_tags]
    # print(f"Product_description : {product_description.text}")
    # print(f"product_tag : {[tag.text for tag in product_tags]}")

    print(data_dict)

    # 이전 버튼 클릭
    back_button = driver.find_element(By.XPATH,
                                      '//*[@id="default-layout"]/div/div[2]/div/div/div/div/button[1]')
    back_button.click()
    time.sleep(3)

    # 엑셀 데이터 입력 [분류, 브랜드이름, 상품이름, 가격, 평점, 성분, 상품설명, 태그]
    ws.cell(row=idx + 1, column=1).value = data_dict["category"]
    ws.cell(row=idx + 1, column=2).value = data_dict["brand_name"]
    ws.cell(row=idx + 1, column=3).value = data_dict["product_name"]
    ws.cell(row=idx + 1, column=4).value = data_dict["price"]
    ws.cell(row=idx + 1, column=5).value = data_dict["rate"]
    ws.cell(row=idx + 1, column=6).value = data_dict["number_reviews"]
    ws.cell(row=idx + 1, column=7).value = str(data_dict["ingredients"])
    ws.cell(row=idx + 1, column=8).value = data_dict["descriptions"]
    ws.cell(row=idx + 1, column=9).value = str(data_dict["tags"])
    wb.save(WORKBOOK_PATH)

# 브라우저 닫기
driver.quit()
